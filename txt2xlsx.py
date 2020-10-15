import openpyxl
from openpyxl.styles import Color,Font,Alignment,PatternFill,Border,Side,Protection
import re

wb = openpyxl.Workbook()
ws = wb.active

ws.cell(row=1, column=1, value='子域名').font = Font(name=u'微软雅黑', bold=True, size=12)
ws.cell(row=1, column=2, value='真实IP').font = Font(name=u'微软雅黑', bold=True, size=12)
ws.cell(row=1, column=3, value='端口').font = Font(name=u'微软雅黑', bold=True, size=12)
ws.cell(row=1, column=4, value='云服务器').font = Font(name=u'微软雅黑', bold=True, size=12)
ws.cell(row=1, column=5, value='C段(不含云服务器)').font = Font(name=u'微软雅黑', bold=True, size=12)
ws.cell(row=1, column=6, value='C段出现次数').font = Font(name=u'微软雅黑', bold=True, size=12)

ws.column_dimensions['A'].width = 32
ws.column_dimensions['B'].width = 18
ws.column_dimensions['C'].width = 32
ws.column_dimensions['D'].width = 44
ws.column_dimensions['E'].width = 22
ws.column_dimensions['F'].width = 13


with open('res.txt', 'r') as f:
    cnt = 2
    for i in f:
        if i:
            r = i.split(', ')
            subdomain = r[0]
            ip = r[1] if re.match(r"^(?:[0-9]{1,3}\.){3}[0-9]{1,3}$", r[1]) else ' '
            ports = r[3].strip().replace(',', '\n') if r[3].strip() != 'NULL' else ' '
            location = r[2]
            ws.cell(row=cnt, column=1, value=subdomain)
            ws.cell(row=cnt, column=2, value=ip)
            ws.cell(row=cnt, column=3, value=ports)
            ws.cell(row=cnt, column=4, value=location)
            cnt += 1

with open('c_ip.txt', 'r') as cip_file:
    cnt = 2
    for i in cip_file:
        if i:
            ws.cell(row=cnt, column=5, value=i.split(' ')[1])
            ws.cell(row=cnt, column=6, value=i.split(' ')[0])
            cnt += 1

for i in range(1, ws.max_row + 1):
    for j in range(1, ws.max_column + 1):
        ws.cell(row=i, column=j).alignment = Alignment(horizontal='center', vertical='center', wrapText=True)  # 水平居中、垂直居中

wb.save('result.xlsx')